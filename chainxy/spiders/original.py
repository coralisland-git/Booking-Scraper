# from __future__ import unicode_literals
import scrapy

import json

import os

import scrapy

from scrapy.spiders import Spider

from scrapy.http import FormRequest

from scrapy.http import Request

from chainxy.items import ChainItem

from scrapy import signals

from scrapy.xlib.pydispatch import dispatcher

from openpyxl import load_workbook

from openpyxl import Workbook 

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

from lxml import etree

from lxml import html

import pdb

import itertools

from operator import itemgetter



class booking_csv(scrapy.Spider):

	name = 'booking_csv'

	domain = 'https://www.booking.com'

	history = []

	global_data = []


	def __init__(self):

		dispatcher.connect(self.spider_closed, signals.spider_closed)

		self.myfile = 'res.xlsx'

		self.mybook = Workbook()

		os.remove(self.myfile) if os.path.exists(self.myfile) else None

		self.mysheet = self.mybook.active

		self.mysheet.title = 'XXX'

	
	def start_requests(self):

		url = "https://www.booking.com/"

		yield scrapy.Request(url, callback=self.parse)


	def parse(self, response):

		checkin_year = "2018"

		checkout_year = "2018"

		checkin_month = "12"

		checkout_month = "12"

		checkin_monthday =15 

		checkout_monthday = 20

		self.duration = checkout_monthday - checkin_monthday

		for day in range(0, self.duration):

			start_date = "2018-12-" + str(checkin_monthday + day)

			end_date = "2018-12-" + str(checkin_monthday + day + 1 )

			url =  "https://www.booking.com/searchresults.en-gb.html?label=gen173nr-1DCAEoggJCAlhYSDNYBGhpiAEBmAEuwgEKd2luZG93cyAxMMgBDNgBA-gBAZICAXmoAgM&lang=en-gb&\
					sid=d27e60ca391f3ed6741fad013f129529&sb=1&src=index&src_elem=sb&\
					error_url=https%3A%2F%2Fwww.booking.com%2Findex.en-gb.html%3Flabel%3Dgen173nr-1DCAEoggJCAlhYSDNYBGhpiAEBmAEuwgEKd2luZG93cyAxMMgBDNgBA-gBAZICAXmoAgM%3Bsid%3Dd27e60ca391f3ed6741fad013f129529%3Bsb_price_type%3Dtotal%26%3B&\
					ss=Bologna&ssne=Bologna&ssne_untouched=Bologna&dest_id=-111742&dest_type=city&\
					checkin_monthday="+start_date+"&checkin_month="+checkin_month+"&checkin_year="+checkin_year+"&checkout_monthday="+end_date+"&checkout_month="+checkout_month+"&checkout_year="+checkout_year+"&no_rooms=1&group_adults=2&group_children=0&b_h4u_keep_filters=&from_sf=1"

			yield scrapy.Request(url=url ,callback=self.parse_pagenation, meta={'date' : start_date})

	def parse_pagenation(self, response):

		count_per_page = 15

		page = 0

		page_count = int(response.xpath('//li[@class="bui-pagination__item sr_pagination_item"]//a//text()').extract()[-1])

		for page in range(0, page_count+1):

			pagenation = "&rows=15&offset="+str(count_per_page * page)

			url = response.url + pagenation

			yield scrapy.Request(url, callback=self.parse_link, meta={'date' : response.meta['date']})


	def parse_link(self, response):

		headers = {
			"Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
			"Accept-Encoding": "gzip, deflate, br",
			"Cookie": "lang_signup_prompt=1; cors_js=1; cws=true; header_signin_prompt=1; _ga=GA1.2.2138684868.1538571570; _gid=GA1.2.844453646.1538571570; BJS=-; has_preloaded=1; _gcl_au=1.1.94462300.1538571615; zz_cook_tms_seg1=1; zz_cook_tms_ed=1; zz_cook_tms_seg3=7; cto_lwid=9fe5eb6f-3617-4549-9d3d-3056eb885bc4; 11_srd=%7B%22features%22%3A%5B%7B%22id%22%3A12%7D%2C%7B%22id%22%3A9%7D%5D%2C%22score%22%3A6%2C%22detected%22%3Afalse%7D; b=%7B%22langPrompt%22%3A%22dontshow%22%7D; lang_signup_prompt=1; bs=%7B%22mtv_user_viewed_comparison_component%22%3Atrue%7D; vpmss=1; _tq_id.TV-81365463-1.3b4c=d292b0bfc219c3c0.1538571626.0.1538627162..; _gat=1; lastSeen=0; utag_main=v_id:01663a03d916007207eb8880f60003073020206b0086e$_sn:8$_ss:0$_st:1538629120029$4split:0$4split2:3$ses_id:1538625489817%3Bexp-session$_pn:14%3Bexp-session; zz_cook_tms_hlist=1971275; bkng=11UmFuZG9tSVYkc2RlIyh9YfDNyGw8J7nzPnUG3Pr%2Bfv5iyz1NtopREN5ChRusm1nwbZQUcG4lajLvU9pK%2BqgWx%2FAunlrUbYEsKO%2FhO6VLnl262%2Fp%2BQQZJcfZ4LmySxgOIBYOMpnBfF%2B3RdcLZWs%2FvdEeaVOvOVRUasA2HCCKjjfIWYE3tmdyHp%2FvDCSYohIUpmhxNAcBN8KEknWOUa%2Fcv%2Bw%3D%3D",
			"Host": "www.booking.com",
			"Upgrade-Insecure-Requests": "1",
			"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36"
		}

		link_list = response.xpath('//a[@class="hotel_name_link url"]//@href').extract()

		for link in link_list:

			link = self.domain + link.strip()

			yield scrapy.Request(link, callback=self.parse_detail, headers=headers, meta={'date' : response.meta['date']})


	def parse_detail(self, response):

		date = response.meta['date']

		estate_name = self.validate(''.join(response.xpath('//h2[@class="hp__hotel-name"]//text()').extract()))

		select_list = response.xpath('//select[@class="hprt-nos-select"]')

		try:

			price_list = self.eliminate_space(select_list[0].xpath('./option//text()').extract())

			res = []

			for price in price_list[1:]:

				price = price.split('(')[1][:-1]

				res.append(price)


			item = {
				"name" : estate_name,
				"date" : date,
				"price" : res
			}

			self.global_data.append(item)

			print('~~~~~~~~~~~~~~~~~', res)

		except Exception as e:

			print('#############', e)

			pass


	def spider_closed(self, spider):

		checkin_monthday =15 

		global_row = 1

		self.global_col = 2

		header_font = Font( b = True, color = 'FF000000' )

		self.mysheet.cell( column = 1 , row = 1,  value = "Name" ).font = header_font

		for col in range(0, self.duration):

			self.mysheet.cell( column = col + self.global_col , row = 1,  value = "2018-12-" + str(checkin_monthday + col) ).font = header_font

		global_row += 1

		global_col = 1

		data = sorted(self.global_data, key=itemgetter('name'))

		for key, group in itertools.groupby(data , key=lambda x:x['name']):

			max_local_row = 0

			for item in list(group):

				local_row_count = len(item['price'])

				local_row = global_row

				for price in item['price']:

					self.mysheet.cell( column = self.get_col(item['date']) , row = local_row,  value = price )

					local_row += 1

				if max_local_row < local_row_count:

					max_local_row_count = local_row_count

			for row in range(0, max_local_row_count):

				self.mysheet.cell( column = 1 , row = global_row + row ,  value = key + '-' + str(row+1) )

			global_row += max_local_row_count

		self.mybook.save( filename = self.myfile )


		try:

			with open('result.json', 'w') as outfile:

				json.dump(self.global_data, outfile)

		except Exception as e:

			print('@@@@@@@@@@', e)


	def get_col(self, date):

		for col_idx in range(0, self.duration):

			if date == self.mysheet.cell( column = col_idx + 2 , row = 1 ).value:

				return col_idx+2

				break

		return 1

	def validate(self, item):

		try:

			return item.replace('\n', '').replace('\t','').replace('\r', '').encode('ascii','ignore').strip()

		except:

			pass


	def eliminate_space(self, items):

	    tmp = []

	    for item in items:

	        if self.validate(item) != '':

	            tmp.append(self.validate(item))

	    return tmp



# "https://www.booking.com/hotel/it/cosmopolitan-bologna.en-gb.html?label=gen173nr-1BCAEoggJCAlhYSDNYBGhpiAEBmAEuwgEKd2luZG93cyAxMMgBDNgBAegBAZICAXmoAgM;sid=0f4fc2416d0f30490cd06bae84692dac;dest_id=-111742;dest_type=city;dist=0;hapos=1;hpos=1;room1=A%2CA;sb_price_type=total;srepoch=1538632420;srfid=caf81ba5ad6036fe01e1bae17fb0a9d19c28d903X1;srpvid=826c2972481d0054;type=total;ucfs=1&#hotelTmpl"
"https://www.booking.com/searchresults.en-gb.html?label=gen173nr-1DCAEoggJCAlhYSDNYBGhpiAEBmAEuwgEKd2luZG93cyAxMMgBDNgBA-gBAZICAXmoAgM&lang=en-gb&%09%09%09%09%09sid=d27e60ca391f3ed6741fad013f129529&sb=1&src=index&src_elem=sb&%09%09%09%09%09error_url=https%3A%2F%2Fwww.booking.com%2Findex.en-gb.html%3Flabel%3Dgen173nr-1DCAEoggJCAlhYSDNYBGhpiAEBmAEuwgEKd2luZG93cyAxMMgBDNgBA-gBAZICAXmoAgM%3Bsid%3Dd27e60ca391f3ed6741fad013f129529%3Bsb_price_type%3Dtotal%26%3B&%09%09%09%09%09ss=Bologna&ssne=Bologna&ssne_untouched=Bologna&dest_id=-111742&dest_type=city&%09%09%09%09%09checkin_monthday=0&checkin_month=12&checkin_year=2018&checkout_monthday=1&checkout_month=12&checkout_year=2018&no_rooms=1&group_adults=2&group_children=0&b_h4u_keep_filters=&from_sf=1"




# latest

# from __future__ import unicode_literals
import scrapy

import json

import os

import scrapy

from scrapy.spiders import Spider

from scrapy.http import FormRequest

from scrapy.http import Request

from chainxy.items import ChainItem

from scrapy import signals

from scrapy.xlib.pydispatch import dispatcher

from openpyxl import load_workbook

from openpyxl import Workbook 

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

from lxml import etree

from lxml import html

import pdb

import itertools

from operator import itemgetter



class booking_latest(scrapy.Spider):

	name = 'booking_latest'

	domain = 'https://www.booking.com'

	history = []

	global_data = []


	def __init__(self):

		dispatcher.connect(self.spider_closed, signals.spider_closed)

		self.myfile = 'res.json'

		os.remove(self.myfile) if os.path.exists(self.myfile) else None

	
	def start_requests(self):

		url = "https://www.booking.com/"

		yield scrapy.Request(url, callback=self.parse)


	def parse(self, response):

		checkin_year = "2018"

		checkout_year = "2018"

		checkin_month = "12"

		checkout_month = "12"

		checkin_monthday =15

		checkout_monthday = 20

		self.duration = checkout_monthday - checkin_monthday

		for day in range(0, self.duration):

			start_date = "2018-12-" + str(checkin_monthday + day)

			end_date = "2018-12-" + str(checkin_monthday + day + 1 )

			url =  "https://www.booking.com/searchresults.en-gb.html?label=gen173nr-1DCAEoggJCAlhYSDNYBGhpiAEBmAEuwgEKd2luZG93cyAxMMgBDNgBA-gBAZICAXmoAgM&lang=en-gb&\
					sid=d27e60ca391f3ed6741fad013f129529&sb=1&src=index&src_elem=sb&\
					error_url=https%3A%2F%2Fwww.booking.com%2Findex.en-gb.html%3Flabel%3Dgen173nr-1DCAEoggJCAlhYSDNYBGhpiAEBmAEuwgEKd2luZG93cyAxMMgBDNgBA-gBAZICAXmoAgM%3Bsid%3Dd27e60ca391f3ed6741fad013f129529%3Bsb_price_type%3Dtotal%26%3B&\
					ss=Bologna&ssne=Bologna&ssne_untouched=Bologna&dest_id=-111742&dest_type=city&\
					checkin_monthday="+start_date+"&checkin_month="+checkin_month+"&checkin_year="+checkin_year+"&checkout_monthday="+end_date+"&checkout_month="+checkout_month+"&checkout_year="+checkout_year+"&no_rooms=1&group_adults=2&group_children=0&b_h4u_keep_filters=&from_sf=1"

			yield scrapy.Request(url=url ,callback=self.parse_pagenation, meta={'date' : start_date})


	def parse_pagenation(self, response):

		count_per_page = 15

		page = 0

		page_count = int(response.xpath('//li[@class="bui-pagination__item sr_pagination_item"]//a//text()').extract()[-1])

		for page in range(0, 1):

			pagenation = "&rows=15&offset="+str(count_per_page * page)

			url = response.url + pagenation

			yield scrapy.Request(url, callback=self.parse_link, meta={'date' : response.meta['date']})


	def parse_link(self, response):

		headers = {
			"Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
			"Accept-Encoding": "gzip, deflate, br",
			"Cookie": "lang_signup_prompt=1; cors_js=1; cws=true; header_signin_prompt=1; _ga=GA1.2.2138684868.1538571570; _gid=GA1.2.844453646.1538571570; BJS=-; has_preloaded=1; _gcl_au=1.1.94462300.1538571615; zz_cook_tms_seg1=1; zz_cook_tms_ed=1; zz_cook_tms_seg3=7; cto_lwid=9fe5eb6f-3617-4549-9d3d-3056eb885bc4; 11_srd=%7B%22features%22%3A%5B%7B%22id%22%3A12%7D%2C%7B%22id%22%3A9%7D%5D%2C%22score%22%3A6%2C%22detected%22%3Afalse%7D; b=%7B%22langPrompt%22%3A%22dontshow%22%7D; lang_signup_prompt=1; bs=%7B%22mtv_user_viewed_comparison_component%22%3Atrue%7D; vpmss=1; _tq_id.TV-81365463-1.3b4c=d292b0bfc219c3c0.1538571626.0.1538627162..; _gat=1; lastSeen=0; utag_main=v_id:01663a03d916007207eb8880f60003073020206b0086e$_sn:8$_ss:0$_st:1538629120029$4split:0$4split2:3$ses_id:1538625489817%3Bexp-session$_pn:14%3Bexp-session; zz_cook_tms_hlist=1971275; bkng=11UmFuZG9tSVYkc2RlIyh9YfDNyGw8J7nzPnUG3Pr%2Bfv5iyz1NtopREN5ChRusm1nwbZQUcG4lajLvU9pK%2BqgWx%2FAunlrUbYEsKO%2FhO6VLnl262%2Fp%2BQQZJcfZ4LmySxgOIBYOMpnBfF%2B3RdcLZWs%2FvdEeaVOvOVRUasA2HCCKjjfIWYE3tmdyHp%2FvDCSYohIUpmhxNAcBN8KEknWOUa%2Fcv%2Bw%3D%3D",
			"Host": "www.booking.com",
			"Upgrade-Insecure-Requests": "1",
			"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36"
		}

		link_list = response.xpath('//a[@class="hotel_name_link url"]//@href').extract()

		for link in link_list[:10]:

			link = self.domain + link.strip()

			"&checkin=2018-12-15&checkout=2018-12-16"
			"&ucf"

			start_date = response.meta['date']

			day = start_date.split('-')[-1]

			end_day = start_date.replace(day, '')  + str(int(day)+1)

			link = link.split("&checkin")[0] + "&checkin=" + response.meta['date'] + "&checkout=" + end_day + "&ucf" + link.split("&ucf")[1]

			yield scrapy.Request(link, callback=self.parse_detail, headers=headers, meta={'date' : response.meta['date']})


	def parse_detail(self, response):

		date = response.meta['date']

		estate_name = self.validate(''.join(response.xpath('//h2[@class="hp__hotel-name"]//text()').extract()))

		select_list = response.xpath('//select[@class="hprt-nos-select"]')

		try:

			price_list = self.eliminate_space(select_list[0].xpath('./option//text()').extract())

			res = []

			idx = 1

			for price in price_list[1:]:

				price = price.split('(')[1][:-1]

				res.append({ idx : price})

				idx += 1

			item = {
				"name" : estate_name,
				"date" : date,
				"price" : res
			}

			self.global_data.append(item)

			print('~~~~~~~~~~~~~~~~~', res)

		except Exception as e:

			print('#############', e)

			pass


	def spider_closed(self, spider):

		data = sorted(self.global_data, key=itemgetter('name'))

		res_data = []

		for key, group in itertools.groupby(data , key=lambda x:x['name']):

			sub_data = []

			for item in list(group):

				sub_data.append(item)

			sub_data = sorted(sub_data, key=itemgetter('date'))

			res_data.append({key : sub_data})

		try:

			with open(self.myfile, 'w') as outfile:

				json.dump(res_data, outfile)

		except Exception as e:

			print('@@@@@@@@@@', e)


	def get_col(self, date):

		for col_idx in range(0, self.duration):

			if date == self.mysheet.cell( column = col_idx + 2 , row = 1 ).value:

				return col_idx+2

				break

		return 1


	def validate(self, item):

		try:

			return item.replace('\n', '').replace('\t','').replace('\r', '').encode('ascii','ignore').strip()

		except:

			pass


	def eliminate_space(self, items):

	    tmp = []

	    for item in items:

	        if self.validate(item) != '':

	            tmp.append(self.validate(item))

	    return tmp




"https://www.booking.com/hotel/it/suite-elite.en-gb.html?label=gen173nr-1DCAEoggJCAlhYSDNYBGhpiAEBmAEuwgEKd2luZG93cyAxMMgBDNgBA-gBAZICAXmoAgM&sid=a385be198292e51a1e7dba73ba8f76f4&checkin=2018-12-15&checkout=2018-12-16&ucfs=1&srpvid=5d948517cc610207&srepoch=1538679344&highlighted_blocks=18734801_102467164_0_1_0&all_sr_blocks=18734801_102467164_0_1_0&room1=A,A&hpos=1&hapos=1&dest_type=city&dest_id=-111742&srfid=7bb83114bb203e2ba293b12db8600211e9597fe9X1&from=searchresults;highlight_room=#hotelTmpl"
"https://www.booking.com/hotel/it/suite-elite.en-gb.html?label=gen173nr-1DCAEoggJCAlhYSDNYBGhpiAEBmAEuwgEKd2luZG93cyAxMMgBDNgBA-gBAZICAXmoAgM&sid=15e2abe919bf545d5e83b1fb6f812e26&checkin=2018-12-15&checkout=2018-12-16&ucfs=1&srpvid=419b84fbcc070222&srepoch=1538679288&highlighted_blocks=18734801_102467164_0_1_0&all_sr_blocks=18734801_102467164_0_1_0&room1=A,A&hpos=1&hapos=1&dest_type=city&dest_id=-111742&srfid=ac0cb58bb2314f904451d7c063f5f3f6831e3292X1&from=searchresults\n;highlight_room=#hotelTmpl"