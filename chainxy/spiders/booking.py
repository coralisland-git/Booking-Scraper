# from __future__ import unicode_literals
import scrapy

import json

import os

import scrapy

from scrapy.spiders import Spider

from scrapy.http import FormRequest

from scrapy.http import Request

from chainxy.items import ChainItem

from scrapy import signals

from scrapy.xlib.pydispatch import dispatcher

from openpyxl import load_workbook

from openpyxl import Workbook 

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

from lxml import etree

from lxml import html

import pdb

import itertools

from operator import itemgetter

import datetime

from datetime import timedelta


class booking(scrapy.Spider):

	name = 'booking'

	domain = 'https://www.booking.com'

	history = []

	global_data = []


	def __init__(self):

		dispatcher.connect(self.spider_closed, signals.spider_closed)

		self.myfile = 'res.json'

		os.remove(self.myfile) if os.path.exists(self.myfile) else None

	
	def start_requests(self):

		url = "https://www.booking.com/"

		yield scrapy.Request(url, callback=self.parse)


	def parse(self, response):

		url = raw_input(' Location : ')

		checkin_date = self.getStr(url, 'checkin_year=', '&') + '-' + self.getStr(url, 'checkin_month=', '&') + '-' + self.getStr(url, 'checkin_monthday=', '&')

		checkout_date = self.getStr(url, 'checkout_year=', '&') + '-' + self.getStr(url, 'checkout_month=', '&') + '-' + self.getStr(url, 'checkout_monthday=', '&')

		input_start_date = checkin_date.split("-")

		input_end_date = checkout_date.split("-")

		self.date = '-'.join(input_start_date)

		self.duration = (datetime.date(int(input_end_date[0]), int(input_end_date[1]), int(input_end_date[2])) - datetime.date(int(input_start_date[0]), int(input_start_date[1]), int(input_start_date[2]))).days

		next_date = (datetime.date(int(input_start_date[0]), int(input_start_date[1]), int(input_start_date[2])) + timedelta(days=1)).strftime('%Y-%m-%d').split('-')

		# url =  "https://www.booking.com/searchresults.en-gb.html?label=gen173nr-1DCAEoggJCAlhYSDNYBGhpiAEBmAEuwgEKd2luZG93cyAxMMgBDNgBA-gBAZICAXmoAgM&lang=en-gb&\
		# 		sid=d27e60ca391f3ed6741fad013f129529&sb=1&src=index&src_elem=sb&\
		# 		error_url=https%3A%2F%2Fwww.booking.com%2Findex.en-gb.html%3Flabel%3Dgen173nr-1DCAEoggJCAlhYSDNYBGhpiAEBmAEuwgEKd2luZG93cyAxMMgBDNgBA-gBAZICAXmoAgM%3Bsid%3Dd27e60ca391f3ed6741fad013f129529%3Bsb_price_type%3Dtotal%26%3B&\
		# 		ss=Bologna&ssne=Bologna&ssne_untouched=Bologna&dest_id=-111742&dest_type=city&\
		# 		checkin_monthday="+input_start_date[2]+"&checkin_month="+input_start_date[1]+"&checkin_year="+input_start_date[0]+"&checkout_monthday="+next_date[2]+"&checkout_month="+next_date[1]+"&checkout_year="+next_date[0]+"&no_rooms=1&group_adults=2&group_children=0&b_h4u_keep_filters=&from_sf=1"

		yield scrapy.Request(url=url ,callback=self.parse_pagenation)


	def parse_pagenation(self, response):

		count_per_page = 15

		page_count = int(response.xpath('//li[@class="bui-pagination__item sr_pagination_item"]//a//text()').extract()[-1])

		for page in range(0, page_count + 1):

			pagenation = "&rows=15&offset="+str(count_per_page * page)

			url = response.url + pagenation

			yield scrapy.Request(url, callback=self.parse_link)


	def parse_link(self, response):

		headers = {
			"Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
			"Accept-Encoding": "gzip, deflate, br",
			"Cookie": "lang_signup_prompt=1; cors_js=1; cws=true; header_signin_prompt=1; _ga=GA1.2.2138684868.1538571570; _gid=GA1.2.844453646.1538571570; BJS=-; has_preloaded=1; _gcl_au=1.1.94462300.1538571615; zz_cook_tms_seg1=1; zz_cook_tms_ed=1; zz_cook_tms_seg3=7; cto_lwid=9fe5eb6f-3617-4549-9d3d-3056eb885bc4; 11_srd=%7B%22features%22%3A%5B%7B%22id%22%3A12%7D%2C%7B%22id%22%3A9%7D%5D%2C%22score%22%3A6%2C%22detected%22%3Afalse%7D; b=%7B%22langPrompt%22%3A%22dontshow%22%7D; lang_signup_prompt=1; bs=%7B%22mtv_user_viewed_comparison_component%22%3Atrue%7D; vpmss=1; _tq_id.TV-81365463-1.3b4c=d292b0bfc219c3c0.1538571626.0.1538627162..; _gat=1; lastSeen=0; utag_main=v_id:01663a03d916007207eb8880f60003073020206b0086e$_sn:8$_ss:0$_st:1538629120029$4split:0$4split2:3$ses_id:1538625489817%3Bexp-session$_pn:14%3Bexp-session; zz_cook_tms_hlist=1971275; bkng=11UmFuZG9tSVYkc2RlIyh9YfDNyGw8J7nzPnUG3Pr%2Bfv5iyz1NtopREN5ChRusm1nwbZQUcG4lajLvU9pK%2BqgWx%2FAunlrUbYEsKO%2FhO6VLnl262%2Fp%2BQQZJcfZ4LmySxgOIBYOMpnBfF%2B3RdcLZWs%2FvdEeaVOvOVRUasA2HCCKjjfIWYE3tmdyHp%2FvDCSYohIUpmhxNAcBN8KEknWOUa%2Fcv%2Bw%3D%3D",
			"Host": "www.booking.com",
			"Upgrade-Insecure-Requests": "1",
			"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36"
		}

		link_list = response.xpath('//a[@class="hotel_name_link url"]//@href').extract()

		date = self.date.split('-')

		for link in link_list:

			link = self.domain + link.strip()

			for idx in range(0, self.duration):

				start_day = (datetime.date(int(date[0]), int(date[1]), int(date[2])) + timedelta(days=idx)).strftime('%Y-%m-%d')

				end_day = (datetime.date(int(date[0]), int(date[1]), int(date[2])) + timedelta(days=idx+1)).strftime('%Y-%m-%d')

				link = link.split("&checkin")[0] + "&checkin=" + start_day + "&checkout=" + end_day + "&ucf" + link.split("&ucf")[1]

				yield scrapy.Request(link, callback=self.parse_detail, headers=headers, meta={'date' : start_day})


	def parse_detail(self, response):

		date = response.meta['date']

		estate_name = self.validate(''.join(response.xpath('//h2[@class="hp__hotel-name"]//text()').extract()))

		number_of_stars = ''

		try:

			number_of_stars = response.xpath('//span[@class="hp__hotel_ratings__stars nowrap"]//i/@title').extract_first().split('-')[0]

		except:

			pass

		parking = "No"

		p_check = ''.join(response.xpath('//div[contains(@class, "hp_desc_important_facilities")]//div//@data-name-en').extract())

		if 'parking' in p_check.lower():

			parking = "Yes"

		latitude = response.body.split('booking.env.b_map_center_latitude =')[1].split(';')[0].strip()

		longitude = response.body.split('booking.env.b_map_center_longitude =')[1].split(';')[0].strip()

		geolocation = {

			"latitude" : latitude,

			"longitude" : longitude
		}

		address = response.xpath('//span[contains(@class, "js-hp_address_subtitle")]//text()').extract_first().strip()

		reviews = self.eliminate_space(response.xpath('//div[contains(@class, "hotel_large_photp_score")]//text()').extract())

		score = response.xpath('//div[contains(@class, "hotel_large_photp_score")]//@data-review-score').extract_first()

		number_of_review = 0

		for review in reviews:

			if "review" in review.lower():

				number_of_review = review.split(' ')[0]

		room_type = ''

		try:

			room_list = response.xpath('//table[contains(@class, "hprt-table  hprt-table-long-language")]//tr')

			room_type = self.validate(room_list[1].xpath('.//span[contains(@class, "hprt-roomtype-icon-link")]//text()').extract_first())
		
		except:

			pass

		for r_idx in range(1, len(room_list)):

			room = room_list[r_idx]

			bound = room_list[r_idx-1].xpath('./@class').extract_first()

			if 'hprt-table-last-row' in bound:		

				room_type = self.validate(room.xpath('.//span[contains(@class, "hprt-roomtype-icon-link")]//text()').extract_first())

			try:

				b_check = ''.join(room.xpath('.//li[contains(@class, "hprt-green-condition jq_tooltip")][1]//text()').extract())

				breakfast = "No"

				if 'included' in b_check.lower():

					breakfast = "Yes"

				max_numer_of_people = room.xpath('.//div[contains(@class, "hprt-occupancy-occupancy-info")]//@data-title').extract_first().strip().split(':')[1].strip()

				price_list = self.eliminate_space(room.xpath('.//select[@class="hprt-nos-select"]//option//text()').extract())

				res = []

				idx = 1

				for price in price_list[1:]:

					price = price.split('(')[1][:-1]

					res.append({

							"room" : str(idx),

							"price" : price
						})

					idx += 1

				item = {

					"name" : estate_name,

					"date" : date,

					"Room Type" : room_type,

					"Number of Stars" : number_of_stars,

					"Geolocation" : geolocation,

					"Address" : address,

					"Booking Score" : score,

					"Number of Review" : number_of_review,

					"Max Number of People" : max_numer_of_people,

					"Parking" : parking,

					"Breakfast" : breakfast,

					"Price" : res
				}

				self.global_data.append(item)

			except Exception as e:

				print('#############', e)

				pass


	def spider_closed(self, spider):

		data = sorted(self.global_data, key=itemgetter('name'))

		res_data = []

		for key, group in itertools.groupby(data , key=lambda x:x['name']):

			sub_data = []

			for item in list(group):

				sub_data.append(item)

			sub_data = sorted(sub_data, key=itemgetter('date'))

			res_data.append({key : sub_data})

		try:

			with open(self.myfile, 'w') as outfile:

				json.dump(res_data, outfile)

		except Exception as e:

			pass


	def get_col(self, date):

		for col_idx in range(0, self.duration):

			if date == self.mysheet.cell( column = col_idx + 2 , row = 1 ).value:

				return col_idx + 2

				break

		return 1


	def getStr(self, item, start, end):

		try:

			return self.validate(item.split(start)[1].split(end)[0])

		except Exception as e:

			print e
			return ''




	def validate(self, item):

		try:

			return item.replace('\n', '').replace('\t','').replace('\r', '').encode('ascii','ignore').strip()

		except:

			pass


	def eliminate_space(self, items):

	    tmp = []

	    for item in items:

	        if self.validate(item) != '':

	            tmp.append(self.validate(item))

	    return tmp